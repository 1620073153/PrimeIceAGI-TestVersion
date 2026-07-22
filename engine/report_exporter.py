from __future__ import annotations

from pathlib import Path
from typing import Sequence

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from engine.batch_models import BatchEvalResult, InterceptType
from engine.batch_statistics import BatchStatistics, compute_statistics, format_summary


HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_header(ws, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def _auto_width(ws, max_width: int = 50):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, min(len(val), max_width))
        ws.column_dimensions[col_letter].width = max_len + 2


def _write_summary_sheet(wb: openpyxl.Workbook, stats: BatchStatistics):
    ws = wb.active
    ws.title = "summary"
    summary = format_summary(stats)

    headers = ["指标", "值"]
    ws.append(headers)
    _style_header(ws, 2)

    label_map = {
        "mode": "评估模式",
        "total_cases": "总用例数",
        "black_total": "黑样本数",
        "white_total": "白样本数",
        "accuracy": "准确率",
        "false_positive_rate": "误报率",
        "guardrail_intercept_rate": "护栏拦截率",
        "model_intercept_rate": "模型拒答率",
        "guardrail_miss_rate": "护栏漏报率",
        "intercept_rate": "拦截率",
        "miss_rate": "漏报率",
    }

    for key, value in summary.items():
        label = label_map.get(key, key)
        if isinstance(value, float):
            ws.append([label, f"{value:.2%}"])
        else:
            ws.append([label, value])

    # 分类别统计摘要
    if stats.by_category:
        ws.append([])
        ws.append(["分类别统计"])
        cat_headers = ["类别", "用例数", "拦截率", "漏报率", "准确率"]
        ws.append(cat_headers)
        cat_header_row = ws.max_row
        for col in range(1, len(cat_headers) + 1):
            cell = ws.cell(row=cat_header_row, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN

        for key, cat in sorted(stats.by_category.items(), key=lambda x: x[1].total, reverse=True):
            if stats.mode == "guardrail":
                intercept_rate = cat.guardrail_intercept_rate
                miss_rate = cat.guardrail_miss_rate
            else:
                intercept_rate = cat.combined_intercept_rate
                miss_rate = cat.miss_rate
            display_name = cat.category_name or cat.category or key
            ws.append([
                display_name,
                cat.total,
                f"{intercept_rate:.1%}",
                f"{miss_rate:.1%}",
                f"{cat.accuracy:.1%}",
            ])

    _auto_width(ws)


def _write_category_sheet(wb: openpyxl.Workbook, stats: BatchStatistics):
    ws = wb.create_sheet("category_breakdown")

    if stats.mode == "guardrail":
        headers = ["类别", "类别名称", "黑样本数", "白样本数",
                   "护栏拦截率", "模型拒答率", "漏报率", "误报率", "准确率"]
    else:
        headers = ["类别", "类别名称", "黑样本数", "白样本数",
                   "拦截率", "漏报率", "误报率", "准确率"]

    ws.append(headers)
    _style_header(ws, len(headers))

    for key, cat in sorted(stats.by_category.items()):
        if stats.mode == "guardrail":
            ws.append([
                cat.category or "-",
                cat.category_name or "-",
                cat.black_total,
                cat.white_total,
                f"{cat.guardrail_intercept_rate:.2%}",
                f"{cat.model_intercept_rate:.2%}",
                f"{cat.guardrail_miss_rate:.2%}",
                f"{cat.false_positive_rate:.2%}",
                f"{cat.accuracy:.2%}",
            ])
        else:
            ws.append([
                cat.category or "-",
                cat.category_name or "-",
                cat.black_total,
                cat.white_total,
                f"{cat.combined_intercept_rate:.2%}",
                f"{cat.miss_rate:.2%}",
                f"{cat.false_positive_rate:.2%}",
                f"{cat.accuracy:.2%}",
            ])

    _auto_width(ws)


def _write_results_sheet(wb: openpyxl.Workbook, results: Sequence[BatchEvalResult]):
    ws = wb.create_sheet("results")
    headers = [
        "用例编号", "类别", "类别名称", "子类型", "绕过手法", "预期标签", "测试内容",
        "模型回复", "判定结果", "是否符合预期", "裁判理由", "置信度",
        "耗时(ms)", "重复序号", "错误信息",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    for r in results:
        response_preview = r.response_text[:500] if r.response_text else ""
        prompt_preview = r.prompt_text[:300] if r.prompt_text else ""
        ws.append([
            r.case_id,
            r.category or "-",
            r.category_name or "-",
            r.subcategory or "-",
            r.bypass_technique or "-",
            r.expected_label,
            prompt_preview,
            response_preview,
            r.intercept_type.value,
            "是" if r.is_correct else "否",
            r.reason,
            round(r.confidence, 3),
            r.elapsed_ms,
            r.repeat_index,
            r.error or "",
        ])

    _auto_width(ws, max_width=60)


def _write_review_sheet(wb: openpyxl.Workbook, results: Sequence[BatchEvalResult]):
    ws = wb.create_sheet("review")
    review_items = [r for r in results if r.review_required]

    headers = [
        "用例编号", "类别", "预期标签", "测试内容",
        "模型回复", "判定结果", "裁判理由", "置信度", "错误信息",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    for r in review_items:
        response_preview = r.response_text[:500] if r.response_text else ""
        prompt_preview = r.prompt_text[:300] if r.prompt_text else ""
        ws.append([
            r.case_id,
            r.category or "-",
            r.expected_label,
            prompt_preview,
            response_preview,
            r.intercept_type.value,
            r.reason,
            round(r.confidence, 3),
            r.error or "",
        ])

    _auto_width(ws, max_width=60)


def _write_failures_sheet(wb: openpyxl.Workbook, results: Sequence[BatchEvalResult]):
    """只筛出未拦截的黑样本，方便定位漏报"""
    ws = wb.create_sheet("failures")
    failures = [r for r in results if r.expected_label == "block" and r.intercept_type == InterceptType.NOT_BLOCKED]

    headers = [
        "用例编号", "类别名称", "子类型", "绕过手法", "测试内容",
        "模型回复", "裁判理由", "置信度",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    for r in failures:
        response_preview = r.response_text[:500] if r.response_text else ""
        prompt_preview = r.prompt_text[:300] if r.prompt_text else ""
        ws.append([
            r.case_id,
            r.category_name or r.category or "-",
            r.subcategory or "-",
            r.bypass_technique or "-",
            prompt_preview,
            response_preview,
            r.reason,
            round(r.confidence, 3),
        ])

    _auto_width(ws, max_width=60)


def export_report(
    results: Sequence[BatchEvalResult],
    mode: str,
    output_path: str,
) -> str:
    stats = compute_statistics(results, mode)
    wb = openpyxl.Workbook()

    _write_summary_sheet(wb, stats)
    _write_category_sheet(wb, stats)
    _write_results_sheet(wb, results)
    _write_review_sheet(wb, results)
    _write_failures_sheet(wb, results)

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    wb.close()

    return str(path.resolve())
