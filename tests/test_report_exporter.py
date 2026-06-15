from pathlib import Path

from openpyxl import load_workbook

from engine.batch_models import BatchEvalResult, InterceptType
from engine.report_exporter import export_batch_report


def test_export_batch_report_writes_results_summary_and_review_sheets(tmp_path: Path):
    output_file = tmp_path / "batch-report.xlsx"
    results = [
        BatchEvalResult(
            case_id="1",
            prompt_text="普通问题",
            response_text="正常回答",
            category="A",
            subcategory="A1",
            source_file="cases.csv",
            regex_labels=[],
            intercept_type=InterceptType.NOT_BLOCKED,
            judge_reason="模型正常输出",
            judge_confidence=0.93,
            success=True,
            retry_count=0,
            latency_ms=520,
            raw_error=None,
            review_required=False,
        ),
        BatchEvalResult(
            case_id="2",
            prompt_text="危险问题",
            response_text="为了安全合规，本次内容无法展示。",
            category="B",
            subcategory="B1",
            source_file="cases.csv",
            regex_labels=["guardrail"],
            intercept_type=InterceptType.GUARDRAIL_BLOCK,
            judge_reason="命中外部护栏模板",
            judge_confidence=0.86,
            success=True,
            retry_count=1,
            latency_ms=810,
            raw_error=None,
            review_required=True,
        ),
    ]

    summary = {
        "total_cases": 2,
        "processed_cases": 2,
        "skipped_cases": 0,
        "review_required_cases": 1,
        "intercept_counts": {
            "not_blocked": 1,
            "guardrail_block": 1,
        },
    }

    export_batch_report(results=results, summary=summary, output_file=output_file)

    workbook = load_workbook(output_file)
    assert workbook.sheetnames == ["results", "summary", "review"]

    expected_headers = [
        "用例编号",
        "测试内容",
        "模型回复",
        "类别名称",
        "子类型",
        "来源文件",
        "正则标签",
        "拦截结论",
        "裁判理由",
        "裁判置信度",
        "重试次数",
        "耗时(ms)",
        "需人工复核",
    ]

    results_sheet = workbook["results"]
    assert results_sheet.max_row == 3
    assert [cell.value for cell in results_sheet[1]] == expected_headers
    assert results_sheet["A2"].value == "1"
    assert results_sheet["H3"].value == "guardrail_block"
    assert results_sheet["K3"].value == 1
    assert results_sheet["L3"].value == 810
    assert results_sheet["M3"].value is True

    summary_sheet = workbook["summary"]
    assert [cell.value for cell in summary_sheet[1]] == ["指标", "值"]
    summary_rows = {
        summary_sheet[f"A{row}"].value: summary_sheet[f"B{row}"].value
        for row in range(2, summary_sheet.max_row + 1)
    }
    assert summary_rows["总样本数"] == 2
    assert summary_rows["已处理样本数"] == 2
    assert summary_rows["跳过样本数"] == 0
    assert summary_rows["需人工复核数"] == 1
    assert summary_rows["拦截统计-not_blocked"] == 1
    assert summary_rows["拦截统计-guardrail_block"] == 1

    review_sheet = workbook["review"]
    assert review_sheet.max_row == 2
    assert [cell.value for cell in review_sheet[1]] == expected_headers
    assert review_sheet["A2"].value == "2"
    assert review_sheet["H2"].value == "guardrail_block"
    assert review_sheet["K2"].value == 1
    assert review_sheet["L2"].value == 810
    assert review_sheet["M2"].value is True
